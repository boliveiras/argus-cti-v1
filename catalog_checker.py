#!/usr/bin/env python3
"""
catalog_checker.py
──────────────────
Utilitário usado pelo script de geração do boletim diário para:

  1. Ler o catálogo Excel e retornar o conjunto de títulos e fontes já
     catalogados (check_catalog).

  2. Determinar se uma notícia nova é duplicata de algo já registrado
     (is_duplicate).

  2b. Determinar se uma notícia deve ser ignorada por relevância
      organizacional (is_irrelevant) — coluna "Relevância" do catálogo.

  3. Acrescentar novas notícias ao catálogo após a geração do boletim
     (append_to_catalog).

Uso pelo boletim:
    from catalog_checker import check_catalog, is_duplicate, is_irrelevant, append_to_catalog

    known = check_catalog(CATALOG_PATH)
    candidatas = [n for n in candidatas if not is_irrelevant(n, known)[0]]
    noticias_filtradas = [n for n in candidatas if not is_duplicate(n, known)[0]]
    # … gera PDF …
    append_to_catalog(CATALOG_PATH, noticias_novas, data_boletim)
"""

import os
import re
import unicodedata
from datetime import date
from typing import TypedDict

try:
    import openpyxl
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
except ImportError:
    raise SystemExit("openpyxl não instalado. Execute: pip install openpyxl --break-system-packages") from None


# ── Constantes de caminho ─────────────────────────────────────────────────────
CATALOG_PATH = os.path.join(
    os.path.dirname(__file__), "catalog", "catalogo-cti.xlsx"
)

# Colunas do catálogo (1-based, letras para referência)
# A=1  B=2  C=3  D=4  E=5  F=6  G=7  H=8  I=9  J=10  K=11  L=12  M=13  N=14
COL_NUM    = 1
COL_TIPO   = 2
COL_TITULO = 3
COL_RESUMO = 4
COL_PROT   = 5
COL_IMP    = 6
COL_ACOES  = 7
COL_DATANOT= 8
COL_DATACHK= 9
COL_SLA    = 10
COL_FONTE  = 11
COL_CRIT   = 12  # Criticidade: CRÍTICO | ALTO | MÉDIO | BAIXO | INFO
COL_TECH   = 13  # M — Tecnologia/Item (preenchida automaticamente ao catalogar)
COL_RELEV  = 14  # N — Relevância: "Nos afeta" | "Não nos afeta" (preenchida pelo usuário)

HEADER_TECH   = "Tecnologia/Item"
HEADER_RELEV  = "Relevância"
RELEV_OPTIONS = ("Nos afeta", "Não nos afeta")

# Mapeamento criticidade → cor de fundo (hex Excel)
CRIT_COLORS = {
    "CRÍTICO": ("1A1A1A", "FFFFFF"),  # preto  / texto branco
    "ALTO":    ("C0392B", "FFFFFF"),  # vermelho / texto branco
    "MÉDIO":   ("D68910", "FFFFFF"),  # laranja / texto branco
    "BAIXO":   ("F4D03F", "1A1A1A"),  # amarelo / texto preto
    "INFO":    ("2E86AB", "FFFFFF"),  # azul claro / texto branco
}

SHEET_NAME = "Catálogo CTI"


# ── Normalização para comparação ──────────────────────────────────────────────
def _normalize(text: str) -> str:
    """Remove acentos, pontuação e caixa para comparação."""
    if not text:
        return ""
    nfkd = unicodedata.normalize("NFKD", text)
    ascii_str = nfkd.encode("ascii", "ignore").decode("ascii")
    return re.sub(r"[^a-z0-9\s]", "", ascii_str.lower()).strip()


def _url_key(url: str) -> str:
    """Extrai o path da URL sem query string para comparação."""
    url = url.strip().rstrip("/")
    # remove protocolo e domínio
    url = re.sub(r"^https?://[^/]+", "", url)
    # remove query/fragment
    url = re.sub(r"[?#].*", "", url)
    return url.lower()


# ── Estrutura de retorno ──────────────────────────────────────────────────────
class KnownCatalog(TypedDict):
    titles: set       # títulos normalizados já no catálogo
    urls: set         # paths de URLs já no catálogo
    last_row: int     # última linha com dados (para append)
    last_num: int     # último número de notícia (para numerar as próximas)
    tech_relevance: dict  # tecnologia normalizada → ("NOS_AFETA"|"NAO_AFETA", nome original)


# ── 1. Leitura do catálogo ────────────────────────────────────────────────────
def check_catalog(catalog_path: str = CATALOG_PATH) -> KnownCatalog:
    """
    Lê o catálogo Excel e retorna um KnownCatalog com:
      - titles: set de títulos normalizados (linhas NOTÍCIA)
      - urls:   set de URL-paths já registrados (coluna Fonte)
      - last_row: linha da última entrada de dados
      - last_num: último número de notícia (#)
      - tech_relevance: mapa tecnologia → relevância (colunas M/N)
    """
    result: KnownCatalog = {"titles": set(), "urls": set(), "last_row": 1,
                            "last_num": 0, "tech_relevance": {}}

    if not os.path.exists(catalog_path):
        return result

    try:
        wb = openpyxl.load_workbook(catalog_path, read_only=True, data_only=True)
    except Exception as e:
        print(f"[catalog_checker] Aviso: não foi possível ler catálogo — {e}")
        return result

    ws = wb[SHEET_NAME] if SHEET_NAME in wb.sheetnames else wb.active

    last_data_row = 1
    last_num = 0

    for row in ws.iter_rows(min_row=2, values_only=True):
        tipo  = str(row[COL_TIPO  - 1] or "").strip()
        titulo = str(row[COL_TITULO - 1] or "").strip()
        fonte  = str(row[COL_FONTE  - 1] or "").strip()
        num    = row[COL_NUM - 1]

        # Considera linhas com conteúdo relevante
        if not tipo and not titulo:
            continue

        last_data_row += 1

        if "NOTÍCIA" in tipo or "NOTICIA" in tipo:
            if titulo:
                result["titles"].add(_normalize(titulo))
            if isinstance(num, (int, float)):
                last_num = max(last_num, int(num))

            # Tecnologia + Relevância (colunas M/N — podem não existir em
            # catálogos antigos, por isso o len() defensivo)
            tech  = str(row[COL_TECH  - 1] or "").strip() if len(row) >= COL_TECH  else ""
            relev = str(row[COL_RELEV - 1] or "").strip() if len(row) >= COL_RELEV else ""
            if tech and relev:
                rel_norm = _normalize(relev)
                if rel_norm == "nos afeta":
                    result["tech_relevance"][_normalize(tech)] = ("NOS_AFETA", tech)
                elif rel_norm == "nao nos afeta":
                    result["tech_relevance"][_normalize(tech)] = ("NAO_AFETA", tech)
                # valor não reconhecido → ignorado (fail-open)

        # Extrair todas as URLs da coluna Fonte
        if fonte:
            for url in fonte.split("\n"):
                url = url.strip()
                if url:
                    result["urls"].add(_url_key(url))

    wb.close()
    result["last_row"]  = last_data_row + 2   # +2 para pular separador
    result["last_num"]  = last_num
    return result


# ── 2. Verificação de duplicata ───────────────────────────────────────────────
def is_duplicate(
    news: dict,
    known: KnownCatalog,
    title_threshold: float = 0.75,
) -> tuple[bool, str]:
    """
    Retorna (True, motivo) se a notícia já existe no catálogo.
    Retorna (False, "") caso contrário.

    Critérios (qualquer um é suficiente):
      1. URL exata (path) já registrada na coluna Fonte.
      2. Similaridade de título > title_threshold (Jaccard sobre tokens).

    Exceções — notícia pode reaparecer se:
      - Título contém "[ATUALIZAÇÃO]" ou "ATUALIZAÇÃO" em maiúsculas.
      - Campo "update" == True no dict da notícia.
    """
    # Permitir re-publicação explícita
    titulo = news.get("titulo", "")
    if "[ATUALIZAÇÃO]" in titulo or news.get("update", False):
        return False, ""

    # 1. Checar URLs
    for url in news.get("fontes", []):
        if _url_key(url) in known["urls"]:
            return True, f"URL já catalogada: {url}"

    # 2. Checar similaridade de título
    norm_new = set(_normalize(titulo).split())
    if norm_new:
        for known_title in known["titles"]:
            norm_known = set(known_title.split())
            if not norm_known:
                continue
            inter = norm_new & norm_known
            union = norm_new | norm_known
            jaccard = len(inter) / len(union) if union else 0
            if jaccard >= title_threshold:
                return True, f"Título similar ({jaccard:.0%}): {known_title[:80]}"

    return False, ""


# ── 2b. Filtro de relevância organizacional ──────────────────────────────────
def is_irrelevant(news: dict, known: KnownCatalog) -> tuple[bool, str]:
    """
    Retorna (True, motivo) se a tecnologia da notícia está marcada como
    "Não nos afeta" no catálogo (coluna Relevância). Nesse caso a notícia
    deve ser IGNORADA — inclusive atualizações críticas/[ATUALIZAÇÃO].

    Retorna (False, "") se:
      - a notícia não tem campo "tecnologia" (notícia geral) → incluir;
      - a tecnologia não está classificada no catálogo (fail-open) → incluir;
      - a tecnologia está marcada "Nos afeta" → incluir (fluxo normal).

    Matching (determinístico, nesta ordem):
      1. Igualdade exata da tecnologia normalizada.
      2. Tokens da tecnologia catalogada ⊆ tokens da tecnologia da notícia
         (ex.: catálogo "Fortinet" casa com notícia "Fortinet FortiSandbox").
         Se houver múltiplos matches por subset, "Nos afeta" prevalece
         (fail-open a favor da visibilidade).
    """
    tech_raw = str(news.get("tecnologia", "") or "").strip()
    if not tech_raw:
        return False, ""

    tech_norm = _normalize(tech_raw)
    if not tech_norm:
        return False, ""

    # 1. Match exato
    exact = known.get("tech_relevance", {}).get(tech_norm)
    if exact:
        if exact[0] == "NAO_AFETA":
            return True, f"Tecnologia '{exact[1]}' marcada como 'Não nos afeta' no catálogo"
        return False, ""  # NOS_AFETA explícito

    # 2. Match por subset de tokens (catálogo mais genérico que a notícia)
    news_tokens = set(tech_norm.split())
    subset_verdicts = []
    for cat_norm, (rel, original) in known.get("tech_relevance", {}).items():
        cat_tokens = set(cat_norm.split())
        if cat_tokens and cat_tokens <= news_tokens:
            subset_verdicts.append((rel, original))

    if subset_verdicts:
        if any(rel == "NOS_AFETA" for rel, _ in subset_verdicts):
            return False, ""  # fail-open: qualquer "Nos afeta" prevalece
        _, original = subset_verdicts[0]
        return True, f"Tecnologia '{original}' marcada como 'Não nos afeta' no catálogo"

    return False, ""  # não classificada → incluir


# ── 3a. Garantir colunas de relevância (cabeçalho + dropdown) ────────────────
def _ensure_relevance_setup(ws) -> None:
    """
    Garante que as colunas M (Tecnologia/Item) e N (Relevância) existem:
    cabeçalhos estilizados como os demais e dropdown de validação em N.
    Idempotente — seguro chamar a cada append.
    """
    from copy import copy as _c

    from openpyxl.worksheet.datavalidation import DataValidation

    ref = ws.cell(row=1, column=COL_TITULO)  # célula-modelo do cabeçalho
    for col, texto in ((COL_TECH, HEADER_TECH), (COL_RELEV, HEADER_RELEV)):
        h = ws.cell(row=1, column=col)
        if not h.value:
            h.value = texto
            h.fill = _c(ref.fill); h.font = _c(ref.font)
            h.alignment = _c(ref.alignment); h.border = _c(ref.border)

    letra_m = openpyxl.utils.get_column_letter(COL_TECH)
    letra_n = openpyxl.utils.get_column_letter(COL_RELEV)
    if not ws.column_dimensions[letra_m].width or ws.column_dimensions[letra_m].width < 20:
        ws.column_dimensions[letra_m].width = 22
    if not ws.column_dimensions[letra_n].width or ws.column_dimensions[letra_n].width < 14:
        ws.column_dimensions[letra_n].width = 16

    # Dropdown "Nos afeta / Não nos afeta" (só adiciona se ainda não existe)
    formula = f'"{",".join(RELEV_OPTIONS)}"'
    ja_existe = any(
        dv.formula1 == formula for dv in ws.data_validations.dataValidation
    )
    if not ja_existe:
        dv = DataValidation(type="list", formula1=formula, allow_blank=True,
                            showDropDown=False)
        dv.error = "Use apenas: Nos afeta / Não nos afeta"
        dv.errorTitle = "Valor inválido"
        ws.add_data_validation(dv)
        dv.add(f"{letra_n}2:{letra_n}5000")



# ── 3b. Criacao do catalogo (primeira execucao standalone) ────────────────---
HEADERS = ("#", "Tipo", "Titulo", "Resumo", "Como se Proteger", "Impacto",
           "Acoes", "Data Noticia", "Data Verificacao", "SLA (h)", "Fonte",
           "Criticidade", HEADER_TECH, HEADER_RELEV)
COL_WIDTHS = (5, 12, 38, 46, 38, 34, 42, 13, 13, 9, 32, 12, 22, 16)


def create_catalog(catalog_path: str) -> None:
    """Cria o catalogo Excel vazio com cabecalhos estilizados (idempotente)."""
    if os.path.exists(catalog_path):
        return
    os.makedirs(os.path.dirname(catalog_path) or ".", exist_ok=True)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = SHEET_NAME
    side = Side(style="thin", color="0D1B2A")
    for col, (texto, largura) in enumerate(zip(HEADERS, COL_WIDTHS, strict=True), start=1):
        c = ws.cell(row=1, column=col, value=texto)
        c.fill = PatternFill("solid", fgColor="0D1B2A")
        c.font = Font(bold=True, color="FFFFFF", size=9, name="Calibri")
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = Border(left=side, right=side, top=side, bottom=side)
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = largura
    ws.row_dimensions[1].height = 28
    ws.freeze_panes = "A2"
    wb.save(catalog_path)
    print(f"[catalog_checker] Catalogo criado: {catalog_path}")


# ── 3. Acrescentar notícias ao catálogo ──────────────────────────────────────
def append_to_catalog(
    catalog_path: str,
    news_list: list[dict],
    data_boletim: date,
    known: KnownCatalog | None = None,
) -> int:
    """
    Acrescenta `news_list` ao catálogo Excel existente.
    Retorna o número de entradas adicionadas.

    Cada item de news_list deve ter:
      titulo, resumo, proteger, impacto, acoes (list[str]), fontes (list[str])
      e opcionalmente tecnologia (str — produto/fornecedor principal)
    """
    if not news_list:
        return 0

    if not os.path.exists(catalog_path):
        create_catalog(catalog_path)

    wb = openpyxl.load_workbook(catalog_path)
    ws = wb[SHEET_NAME] if SHEET_NAME in wb.sheetnames else wb.active

    _ensure_relevance_setup(ws)

    # Encontrar última linha com dados e último número de notícia
    max_row = ws.max_row
    last_num = 0
    for r in range(2, max_row + 1):
        tipo = str(ws.cell(r, COL_TIPO).value or "")
        num  = ws.cell(r, COL_NUM).value
        if ("NOTÍCIA" in tipo or "NOTICIA" in tipo) and isinstance(num, (int, float)):
            last_num = max(last_num, int(num))

    # Se known fornecido, usa o last_num dele (mais eficiente)
    if known and known["last_num"] > last_num:
        last_num = known["last_num"]

    # Paleta (mesma do gerador)
    C_MID_BLUE = "1B3A5C"; C_ACCENT_BLUE = "2E86AB"
    C_ACTION_BG   = "EBF5FB"; C_ACTION_ALT = "E3F2FD"
    C_WHITE = "FFFFFF"; C_TEXT = "1A1A2E"; C_ACTION_TEXT = "1B3A5C"
    C_ORANGE = "D68910"; C_BORDER = "CBD5E1"

    def _fill(hex): return PatternFill("solid", fgColor=hex)
    def _font(bold=False, color=C_TEXT, size=9, italic=False):
        return Font(bold=bold, color=color, size=size, italic=italic, name="Calibri")
    def _align(h="left", v="top", wrap=True):
        return Alignment(horizontal=h, vertical=v, wrap_text=wrap)
    def _border(color=C_BORDER):
        s = Side(style="thin", color=color)
        return Border(left=s, right=s, top=s, bottom=s)

    row = max_row + 1
    added = 0

    for news in news_list:
        last_num += 1
        item_num  = last_num
        is_alt    = (item_num % 2 == 0)
        bg_n      = C_MID_BLUE if not is_alt else "24476E"
        bg_a      = C_ACTION_BG if not is_alt else C_ACTION_ALT
        fonte_txt = "\n".join(news.get("fontes", []))
        acoes_txt = "\n".join(news.get("acoes", []))

        noticia_row = row
        ws.row_dimensions[row].height = 72
        nb = _border("1B3A5C")

        def nc(col, val, bold=False, ha="left", fmt=None,
               noticia_row=noticia_row, bg_n=bg_n, nb=nb):
            c = ws.cell(row=noticia_row, column=col, value=val)
            c.fill = _fill(bg_n)
            c.font = _font(bold=bold, color=C_WHITE)
            c.alignment = _align(h=ha, v="top")
            c.border = nb
            if fmt: c.number_format = fmt
            return c

        nc(COL_NUM,    item_num, bold=True, ha="center").alignment = _align("center","center",False)
        nc(COL_TIPO,   "📋 NOTÍCIA", bold=True, ha="center").alignment = _align("center","center",False)
        nc(COL_TITULO, news.get("titulo",""), bold=True)
        nc(COL_RESUMO, news.get("resumo",""))
        nc(COL_PROT,   news.get("proteger",""))
        nc(COL_IMP,    news.get("impacto",""))
        nc(COL_ACOES,  acoes_txt)
        d = nc(COL_DATANOT, data_boletim, fmt="DD/MM/YYYY")
        d.alignment = _align("center","center",False)
        e = nc(COL_DATACHK, None); e.number_format = "DD/MM/YYYY"
        e.alignment = _align("center","center",False)

        # Criticidade — coluna L com cor própria
        crit_val = news.get("criticidade", "INFO").upper()
        crit_bg, crit_fg = CRIT_COLORS.get(crit_val, ("2E86AB", "FFFFFF"))
        cc = ws.cell(row=noticia_row, column=COL_CRIT, value=crit_val)
        cc.fill = PatternFill("solid", fgColor=crit_bg)
        cc.font = _font(bold=True, color=crit_fg, size=8)
        cc.alignment = _align("center", "center", False)
        cc.border = nb

        cf = ws.cell(row=noticia_row, column=COL_SLA,
                     value=f'=IF(I{noticia_row}="","",(I{noticia_row}-H{noticia_row})*24)')
        cf.fill = _fill(bg_n); cf.font = _font(bold=True, color=C_WHITE)
        cf.alignment = _align("center","center",False); cf.border = nb
        cf.number_format = '0.0"h"'

        ck = ws.cell(row=noticia_row, column=COL_FONTE, value=fonte_txt)
        ck.fill = _fill(bg_n); ck.font = _font(color="A8C8E8", size=8, italic=True)
        ck.alignment = _align(); ck.border = nb

        # Tecnologia/Item (M) — preenchida automaticamente; Relevância (N) —
        # fica em branco para o usuário classificar via dropdown
        ct = nc(COL_TECH, news.get("tecnologia", ""), bold=True, ha="center")
        ct.alignment = _align("center", "center")
        cr = ws.cell(row=noticia_row, column=COL_RELEV, value=None)
        cr.fill = PatternFill("solid", fgColor="FFFDE7")  # amarelo claro = pendente
        cr.alignment = _align("center", "center", False)
        s_p = Side(style="thin", color=C_ORANGE)
        cr.border = Border(left=s_p, right=s_p, top=s_p, bottom=s_p)

        row += 1

        for idx, acao in enumerate(news.get("acoes",[]), 1):
            acao_row = row
            ws.row_dimensions[row].height = 38
            ab = _border()

            def ac(col, val=None, bold=False, ha="left", fmt=None, color=C_ACTION_TEXT,
                   acao_row=acao_row, bg_a=bg_a, ab=ab):
                c = ws.cell(row=acao_row, column=col, value=val)
                c.fill = _fill(bg_a)
                c.font = _font(bold=bold, color=color)
                c.alignment = _align(h=ha, v="center")
                c.border = ab
                if fmt: c.number_format = fmt
                return c

            ac(COL_NUM,    f"↳{idx}", bold=True, ha="center", color=C_ACCENT_BLUE)
            c2 = ws.cell(row=acao_row, column=COL_TIPO, value="  ✅ AÇÃO")
            c2.fill = _fill(bg_a); c2.font = _font(italic=True, color=C_ACCENT_BLUE, size=8)
            c2.alignment = _align("left","center",False); c2.border = ab

            c3 = ws.cell(row=acao_row, column=COL_TITULO, value=news.get("titulo",""))
            c3.fill = _fill(bg_a); c3.font = _font(italic=True, color="4A6785", size=8)
            c3.alignment = _align("left","center"); c3.border = ab

            for col_e in [COL_RESUMO, COL_PROT, COL_IMP]: ac(col_e)
            ac(COL_ACOES, acao)

            ch = ws.cell(row=acao_row, column=COL_DATANOT, value=data_boletim)
            ch.fill = _fill(bg_a); ch.font = _font(color=C_ACTION_TEXT, size=8)
            ch.alignment = _align("center","center",False); ch.border = ab
            ch.number_format = "DD/MM/YYYY"

            ci = ws.cell(row=acao_row, column=COL_DATACHK, value=None)
            ci.fill = PatternFill("solid", fgColor="FFFDE7")
            ci.font = _font(color=C_TEXT)
            ci.alignment = _align("center","center",False)
            ci.number_format = "DD/MM/YYYY"
            s_o = Side(style="thin", color=C_ORANGE)
            ci.border = Border(left=s_o, right=s_o, top=s_o, bottom=s_o)

            cj = ws.cell(row=acao_row, column=COL_SLA,
                         value=f'=IF(I{acao_row}="","",(I{acao_row}-H{acao_row})*24)')
            cj.fill = PatternFill("solid", fgColor="EBF5FB")
            cj.font = _font(bold=True, color=C_MID_BLUE)
            cj.alignment = _align("center","center",False)
            cj.number_format = '0.0"h"'
            s_a = Side(style="thin", color=C_ACCENT_BLUE)
            cj.border = Border(left=s_a, right=s_a, top=s_a, bottom=s_a)

            ck_a = ws.cell(row=acao_row, column=COL_FONTE, value=fonte_txt)
            ck_a.fill = _fill(bg_a); ck_a.font = _font(color="7A9BB5", size=7, italic=True)
            ck_a.alignment = _align("left","center"); ck_a.border = ab

            # Criticidade nas linhas de ação (herdada da notícia)
            cc_a = ws.cell(row=acao_row, column=COL_CRIT, value=crit_val)
            cc_a.fill = PatternFill("solid", fgColor=crit_bg)
            cc_a.font = _font(italic=True, color=crit_fg, size=7)
            cc_a.alignment = _align("center", "center", False)
            cc_a.border = ab

            # Tecnologia herdada + Relevância vazia (estética nas linhas de ação)
            ct_a = ws.cell(row=acao_row, column=COL_TECH, value=news.get("tecnologia", ""))
            ct_a.fill = _fill(bg_a); ct_a.font = _font(italic=True, color="4A6785", size=8)
            ct_a.alignment = _align("center", "center", False); ct_a.border = ab
            cr_a = ws.cell(row=acao_row, column=COL_RELEV, value=None)
            cr_a.fill = _fill(bg_a); cr_a.border = ab

            row += 1

        # Separador
        for col_s in range(1, COL_RELEV + 1):
            cs = ws.cell(row=row, column=col_s)
            cs.fill = PatternFill("solid", fgColor="C8D8E8")
            cs.border = Border(top=Side(style="thin", color="A0B4C8"))
        ws.row_dimensions[row].height = 4
        row += 1
        added += 1

    wb.save(catalog_path)
    print(f"[catalog_checker] {added} notícia(s) adicionada(s) ao catálogo.")
    return added


# ── CLI de teste ──────────────────────────────────────────────────────────────
if __name__ == "__main__":
    known = check_catalog(CATALOG_PATH)
    print(f"Títulos conhecidos : {len(known['titles'])}")
    print(f"URLs conhecidas    : {len(known['urls'])}")
    print(f"Último nº de item  : {known['last_num']}")
    print(f"Última linha       : {known['last_row']}")
    print(f"Techs classificadas: {len(known['tech_relevance'])}")

    # Teste de duplicata
    test = {
        "titulo": "CVE-2026-0625: Zero-Day em Roteadores D-Link Descontinuados",
        "fontes": ["https://www.darkreading.com/cyberattacks-data-breaches/attackers-exploit-zero-day-end-of-life-d-link-routers"],
    }
    dup, motivo = is_duplicate(test, known)
    print(f"\nTeste duplicata '{test['titulo'][:50]}...'")
    print(f"  Duplicata: {dup}  |  Motivo: {motivo or '—'}")

    # Teste de relevância
    test_rel = {"titulo": "Falha no FortiSandbox", "tecnologia": "Fortinet FortiSandbox"}
    irr, motivo = is_irrelevant(test_rel, known)
    print(f"\nTeste relevância '{test_rel['tecnologia']}'")
    print(f"  Irrelevante: {irr}  |  Motivo: {motivo or '—'}")
